import openpyxl
import json
import warnings

# 경고 메시지 무시 설정
warnings.simplefilter(action='ignore', category=UserWarning)

# 파일 경로 설정
excel = r'C:\Users\kmg00\Desktop\생활쓰레기종합배출정보.xlsx'
json_f = r'C:\Users\kmg00\Desktop\Rdata\생활쓰레기종합배출정보.json'

# 액셀 파일 열기
wb = openpyxl.load_workbook(excel, read_only=True)

# 첫 번째 시트 가져오기
sheet = wb.worksheets[0]

# 첫 번째 행(헤더)의 값을 키로 사용하기 위해 key_list 생성
key_list = []
for col in range(1, sheet.max_column + 1):
    key_list.append(sheet.cell(row=1, column=col).value)

# 각 행의 데이터를 딕셔너리로 변환하여 data 리스트에 추가
data = []
for row in range(2, sheet.max_row + 1):
    temp = {}
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=row, column=col).value
        temp[key_list[col-1]] = val
    data.append(temp)

# 액셀 파일 닫기
wb.close()

# 추출된 데이터 출력
print(data)

# 데이터를 JSON 파일로 저장
try:
    with open(json_f, 'w', encoding='utf-8') as fp:
        json.dump(data, fp, indent=4, ensure_ascii=False)
except Exception as e:
    print(f"Error while saving to {json_f}: {e}")

print('성공')
